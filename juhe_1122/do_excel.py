#!/usr/bin/python
#-*- coding:utf-8 -*-
#@Author：zhuxiujie

"""将操作excel封装函数"""
from openpyxl import load_workbook

class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):

        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1,sheet.max_row+1):
            sub_data = {}
            sub_data['method'] = sheet.cell(i,1).value
            sub_data['url'] = sheet.cell(i,2).value
            sub_data['data'] = sheet.cell(i, 3).value
            sub_data['expected'] = sheet.cell(i, 4).value
            test_data.append(sub_data)
        return test_data

if __name__ == '__main__':
    print(DoExcel("data.xlsx","old").get_data())
