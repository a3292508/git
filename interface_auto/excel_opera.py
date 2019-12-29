#!usr/bin/env python
#-*- coding:utf-8 -*-

import os
from openpyxl import load_workbook
from interface_auto.common.read_config import ReadConfig
from interface_auto.common.project_path import data_path

file_name = ReadConfig().read_config('DATA','file_name')
test_data = os.path.join(data_path,file_name)
# print(test_data)

workbook = load_workbook(test_data)
sheet = workbook['login']
# data = sheet.cell(1,1).value
# print(data)
# max_row = sheet.max_row
# max_column = sheet.max_column
# print(max_row)
# print(max_column)
# title_list = []
# for i in range(1,sheet.max_column+1):
#     title_list.append(sheet.cell(1,i).value)
# print(title_list)
data = []
for item in range(2,sheet.max_row+1):
    sub_data = {}
    sub_data['case_id'] = sheet.cell(item,1).value
    sub_data['interface'] = sheet.cell(item,2).value
    sub_data['title'] = sheet.cell(item,3).value
    sub_data['url'] = sheet.cell(item,4).value
    sub_data['data'] = sheet.cell(item,5).value
    sub_data['method'] = sheet.cell(item,6).value
    sub_data['expected'] = sheet.cell(item,7).value
    # sub_data['result'] = sheet.cell(8,item).value
    data.append(sub_data)
for d in data:
    print(d)
