#!usr/bin/env python
#-*- coding:utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook('test_data.xlsx')
sheet = wb['login']

test_data = []
for item in range(2,sheet.max_row+1):
    row_data = {}
    row_data['url'] = sheet.cell(item,1).value
    row_data['data'] = sheet.cell(item, 2).value
    row_data['title'] = sheet.cell(item, 3).value
    row_data['http_method'] = sheet.cell(item, 4).value
    test_data.append(row_data)
print(test_data)