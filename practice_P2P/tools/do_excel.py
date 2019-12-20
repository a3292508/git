#!usr/bin/env python
#-*- coding:utf-8 -*-

from openpyxl import load_workbook
from practice_P2P.tools.read_config import ReadConfig
from practice_P2P.tools import project_path
from practice_P2P.tools.get_data import GetData

class DoExcel:
    """操作excel文件"""

    # @staticmethod     #可以设置成静态函数
    def get_data(self,file_name):
        """获取数据"""
        wb = load_workbook(file_name)
        mode = eval(ReadConfig.get_config(project_path.test_config_path,'MODE','mode'))
        tel = GetData.Telphone      #从GetData里面拿到数据

        test_data = []
        for key in mode:        #遍历存在配置文件里的字典
            sheet = wb[key]
            if mode[key] == 'all':
                for item in range(2,sheet.max_row+1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(item,1).value
                    row_data['url'] = sheet.cell(item,2).value
                    # row_data['data'] = sheet.cell(item,3).value
                    if sheet.cell(item,3).value.find('${tel}')!=-1:
                        row_data['data'] = sheet.cell(item,3).value.replace('${tel}',str(tel))
                    # elif sheet.cell(item,3).value.find('${tel}')!=-1:
                    #     row_data['data'] = sheet.cell(item, 3).value.replace('${tel}', str(tel+1))
                    else:
                        row_data['data'] = sheet.cell(item, 3).value
                    row_data['title'] = sheet.cell(item,4).value
                    row_data['http_method'] = sheet.cell(item,5).value
                    row_data['expected'] = sheet.cell(item,6).value
                    test_data.append(row_data)
                    self.update_tel(file_name,'init',tel+2)
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['url'] = sheet.cell(case_id+1, 2).value
                    # row_data['data'] = sheet.cell(case_id+1, 3).value
                    if sheet.cell(case_id+1,3).value.find('${tel_1}')!=-1:
                        row_data['data'] = sheet.cell(case_id+1,3).value.replace('${tel_1}',str(tel))
                    elif sheet.cell(case_id+1,3).value.find('${tel}')!=-1:
                        row_data['data'] = sheet.cell(case_id+1, 3).value.replace('${tel}', str(tel+1))
                    else:
                        row_data['data'] = sheet.cell(case_id+1, 3).value
                    row_data['title'] = sheet.cell(case_id+1, 4).value
                    row_data['http_method'] = sheet.cell(case_id+1, 5).value
                    row_data['expected'] = sheet.cell(case_id+1, 6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
                    self.update_tel(file_name, 'init', tel + 2)
        return test_data

    @staticmethod
    def write_back(file_name,sheet_name,item,result,test_result):
        """写回结果数据"""
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(item,7).value = result
        sheet.cell(item,8).value = test_result
        wb.save(file_name)

    def update_tel(self,file_name,sheet_name,tel):
        """更新Excel里的手机号数据"""
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2,1).value = tel
        wb.save(file_name)

if __name__ == '__main__':
    test_data = DoExcel().get_data(project_path.test_data_path)
    # for data in test_data:
    #     print(data)
    print(test_data)
