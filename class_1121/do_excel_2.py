#!usr/bin/env python
#-*- coding:utf-8 -*-

"""将操作excel封装函数"""
from openpyxl import load_workbook

class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]     #获取一个表单对象
        self.get_max_row = self.sheet_obj.max_row

    def get_data(self,i,j):
        """根据传入的坐标来获取值"""
        return self.sheet_obj.cell(i,j).value

if __name__ == '__main__':
    print(DoExcel("test.xlsx","practice").get_data(1,1))


#方法一和方法二，读取数据都可以。一般使用方法一，更易于维护和理解
#方法一：一次性读取所有数据，对内存的要求高一些
#方法二：需要用的时候读取所有数据，对磁盘读写要求高一些
#速度：CPU>内存>磁盘