#!/usr/bin/python
#-*- coding:utf-8 -*-
#@Author：zhuxiujie

from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
sheet = wb['practice']
test_data = []
for i in range(1,sheet.max_row+1):
    sub_data = {}
    sub_data['method'] = sheet.cell(i,1).value
    sub_data['url'] = sheet.cell(i,2).value
    sub_data['data'] = sheet.cell(i, 3).value
    sub_data['expected'] = sheet.cell(i, 4).value
    test_data.append(sub_data)
print(test_data)
# for i in range(len(test_data)):
#     print(test_data[i])