#!/usr/bin/python
#-*- coding:utf-8 -*-
#@Author：zhuxiujie

from openpyxl import load_workbook

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_header(self):
        """获取第一行的标题行"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = []     #存储标题行
        for j in range(1,sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self):
        """根据嵌套循环读取数据"""
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = self.get_header()
        test_data = []

        for i in range(2,sheet.max_row+1):
            sub_data = {}
            for j in range(1,sheet.max_column+1):
                sub_data[header[j-1]] = sheet.cell(i,j).value       #j-1是因为header是列表，从下标0开始取值
            test_data.append(sub_data)
        return test_data

if __name__ == '__main__':
    print(DoExcel('python.xlsx', 'python').get_header())
    print(DoExcel('python.xlsx','python').get_data())
