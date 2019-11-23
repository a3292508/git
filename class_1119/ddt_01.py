#!/usr/bin/python
#-*- coding:utf-8 -*-
#@Author：zhuxiujie

#python操作Excel
#地址、测试数据、断言期望结果，除了这几个不同，其他基本相同

#1.只支持这种后缀.xlsx---->openpyxl 只支持这种格式
#2.不能直接在编辑器中新建xlsx文件

from openpyxl import load_workbook

#1.打开excel
wb = load_workbook('test.xlsx')
#2.定位表单
sheet = wb['practice']
#3.定位单元格，行列值
res = sheet.cell(1,1).value

print('拿到的结果是：',res)
print('最大行，{}'.format(sheet.max_row))       #最大行
print('最大列，{}'.format(sheet.max_column))    #最大列

print('结果是：{},类型是{}'.format(sheet.cell(1,1).value,type(sheet.cell(1,1).value)))
print('结果是：{},类型是{}'.format(sheet.cell(1,2).value,type(sheet.cell(1,2).value)))
print('结果是：{},类型是{}'.format(sheet.cell(1,3).value,type(sheet.cell(1,3).value)))
print('结果是：{},类型是{}'.format(sheet.cell(1,4).value,type(sheet.cell(1,4).value)))

#数据从Excel里面拿出来是什么类型：
#数字还是数字，其他类型都是字符串！